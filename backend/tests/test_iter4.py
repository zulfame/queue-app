"""Backend tests iteration 4: recap/export XLSX, surveys, db backup/restore, printer & footer settings."""
import os
import io
import json
import time
import pytest
import requests

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL')
if not BASE_URL:
    with open('/app/frontend/.env') as f:
        for line in f:
            if line.startswith('REACT_APP_BACKEND_URL='):
                BASE_URL = line.split('=', 1)[1].strip()
                break
BASE_URL = BASE_URL.rstrip('/')
API = f"{BASE_URL}/api"

ADMIN = ("admin@antrian.id", "admin123")
OP_PUSAT = ("operator.pusat@antrian.id", "operator123")
OP_CABANG = ("operator.cabang@antrian.id", "operator123")


def login(email, password):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, f"login {email} -> {r.status_code} {r.text}"
    return r.json()


def H(tok):
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def admin_h():
    return H(login(*ADMIN)["access_token"])


@pytest.fixture(scope="module")
def op_pusat_h():
    return H(login(*OP_PUSAT)["access_token"])


@pytest.fixture(scope="module")
def op_cabang_h():
    return H(login(*OP_CABANG)["access_token"])


@pytest.fixture(scope="module")
def branches():
    return requests.get(f"{API}/branches").json()


@pytest.fixture(scope="module")
def branch_pusat(branches):
    return next(b for b in branches if b["name"] == "Kantor Pusat")


@pytest.fixture(scope="module")
def branch_cabang(branches):
    return next(b for b in branches if b["name"] == "Kantor Cabang")


def _svc_in(branch_id):
    services = requests.get(f"{API}/services", params={"branch_id": branch_id}).json()
    assert services, f"no services in branch {branch_id}"
    return services[0]


def _counter_in(branch_id):
    counters = requests.get(f"{API}/counters", params={"branch_id": branch_id}).json()
    assert counters, f"no counters in branch {branch_id}"
    return counters[0]


# ---------- Recap XLSX export ----------
class TestRecapExport:
    def test_export_requires_auth(self):
        r = requests.get(f"{API}/recap/export")
        assert r.status_code == 401

    def test_admin_export_xlsx_valid(self, admin_h, branch_cabang):
        # Create a call log to have some data
        svc = _svc_in(branch_cabang["id"])
        counter = _counter_in(branch_cabang["id"])
        t = requests.post(f"{API}/tickets", json={"service_id": svc["id"], "branch_id": branch_cabang["id"]}).json()
        requests.post(f"{API}/queue/call-next", headers=admin_h,
                      json={"counter_id": counter["id"], "service_id": svc["id"], "branch_id": branch_cabang["id"]})

        r = requests.get(f"{API}/recap/export", headers=admin_h, params={"branch_id": branch_cabang["id"]})
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "officedocument" in ct, ct
        # XLSX signature: PK zip
        assert r.content[:2] == b"PK", r.content[:20]
        # openpyxl can read it
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert "Tiket" in header and "Rating" in header and "Saran" in header

    def test_operator_export_only_own_branch(self, op_cabang_h, branch_pusat, branch_cabang):
        # Even if operator passes different branch_id, backend forces own branch
        r = requests.get(f"{API}/recap/export", headers=op_cabang_h, params={"branch_id": branch_pusat["id"]})
        assert r.status_code == 200
        # We cannot see column contents easily but ensure a valid xlsx returned.
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # All rows (skip header) should have branch_name == cabang or be empty
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            branch_name = row[6]
            if branch_name:
                assert branch_name == branch_cabang["name"], f"operator saw {branch_name}"


# ---------- Surveys ----------
class TestSurveys:
    @pytest.fixture(scope="class")
    def a_ticket(self, admin_h, branch_cabang):
        svc = _svc_in(branch_cabang["id"])
        t = requests.post(f"{API}/tickets", json={"service_id": svc["id"], "branch_id": branch_cabang["id"]}).json()
        return t

    def test_survey_requires_auth(self, a_ticket):
        r = requests.post(f"{API}/surveys", json={"ticket_id": a_ticket["id"], "rating": 5})
        assert r.status_code == 401

    def test_ticket_not_found_404(self, admin_h):
        r = requests.post(f"{API}/surveys", headers=admin_h,
                          json={"ticket_id": "nonexistent-id-xyz", "rating": 5, "feedback": "x"})
        assert r.status_code == 404

    def test_submit_survey_and_appears_in_recap(self, admin_h, a_ticket, branch_cabang):
        r = requests.post(f"{API}/surveys", headers=admin_h, json={
            "ticket_id": a_ticket["id"], "rating": 4, "feedback": "Bagus", "photo": ""
        })
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["ok"] is True
        assert d["survey"]["rating"] == 4
        assert d["survey"]["feedback"] == "Bagus"

        # Recap should include this survey attached to logs for that ticket
        rr = requests.get(f"{API}/recap", headers=admin_h, params={"branch_id": branch_cabang["id"]})
        assert rr.status_code == 200
        logs = rr.json()["logs"]
        matching = [l for l in logs if l["ticket_id"] == a_ticket["id"]]
        # There may or may not be logs for this ticket (we didn't necessarily call it).
        # If there are, survey should be attached
        for l in matching:
            assert l.get("survey") is not None
            assert l["survey"]["rating"] == 4

    def test_rating_clamped_high(self, admin_h, a_ticket):
        r = requests.post(f"{API}/surveys", headers=admin_h,
                          json={"ticket_id": a_ticket["id"], "rating": 99, "feedback": ""})
        assert r.status_code == 200
        assert r.json()["survey"]["rating"] == 5

    def test_rating_clamped_low(self, admin_h, a_ticket):
        r = requests.post(f"{API}/surveys", headers=admin_h,
                          json={"ticket_id": a_ticket["id"], "rating": -3, "feedback": ""})
        assert r.status_code == 200
        assert r.json()["survey"]["rating"] == 0

    def test_operator_cross_branch_survey_forbidden(self, op_cabang_h, admin_h, branch_pusat):
        svc = _svc_in(branch_pusat["id"])
        t = requests.post(f"{API}/tickets", json={"service_id": svc["id"], "branch_id": branch_pusat["id"]}).json()
        r = requests.post(f"{API}/surveys", headers=op_cabang_h,
                          json={"ticket_id": t["id"], "rating": 5, "feedback": ""})
        assert r.status_code == 403


# ---------- DB backup / restore ----------
class TestDBBackupRestore:
    def test_backup_requires_admin(self, op_pusat_h):
        r = requests.get(f"{API}/db/backup")
        assert r.status_code == 401
        r2 = requests.get(f"{API}/db/backup", headers=op_pusat_h)
        assert r2.status_code == 403

    def test_restore_requires_admin(self, op_pusat_h):
        r = requests.post(f"{API}/db/restore", json={"data": {}})
        assert r.status_code == 401
        r2 = requests.post(f"{API}/db/restore", headers=op_pusat_h, json={"data": {}})
        assert r2.status_code == 403

    def test_backup_and_restore_roundtrip(self, admin_h):
        # Backup
        r = requests.get(f"{API}/db/backup", headers=admin_h)
        assert r.status_code == 200
        dump = r.json()
        assert "data" in dump
        assert "branches" in dump["data"]
        assert "users" in dump["data"]
        assert "settings" in dump["data"]
        assert len(dump["data"]["branches"]) >= 2
        # ensure no _id in dump
        for coll, docs in dump["data"].items():
            for d in docs:
                assert "_id" not in d, f"{coll} has _id"

        # Add a marker service, then restore original -> marker should disappear
        marker_svc = requests.post(f"{API}/services", headers=admin_h, json={
            "name": "TEST_MARKER_ITER4", "prefix": "ZZ", "description": "", "icon": "users", "active": True
        }).json()
        assert marker_svc["name"] == "TEST_MARKER_ITER4"

        # Restore with dump (which does NOT contain marker)
        rr = requests.post(f"{API}/db/restore", headers=admin_h, json={"data": dump["data"]})
        assert rr.status_code == 200, rr.text
        rd = rr.json()
        assert rd["ok"] is True
        assert "branches" in rd["restored"]
        assert rd["restored"]["branches"] == len(dump["data"]["branches"])

        # Verify marker gone
        services_after = requests.get(f"{API}/services").json()
        names = [s["name"] for s in services_after]
        assert "TEST_MARKER_ITER4" not in names, "restore did not remove marker"

        # Verify branches survived
        branches_after = requests.get(f"{API}/branches").json()
        assert len(branches_after) == len(dump["data"]["branches"])


# ---------- Branch printer & footer settings ----------
class TestPrinterAndFooter:
    def test_put_branch_printer_fields(self, admin_h, branch_cabang):
        payload = {
            "name": branch_cabang["name"],
            "address": branch_cabang.get("address", ""),
            "active": True,
            "ticker_text": branch_cabang.get("ticker_text", ""),
            "promo_media": branch_cabang.get("promo_media", []),
            "printer_name": "TEST_PRINTER_A",
            "print_header": "TEST_HEADER_TXT",
            "print_footer": "TEST_FOOTER_TXT",
        }
        r = requests.put(f"{API}/branches/{branch_cabang['id']}", headers=admin_h, json=payload)
        assert r.status_code == 200, r.text
        # PUT returns {ok:True}; verify persistence via GET
        b_all = requests.get(f"{API}/branches").json()
        bc = next(x for x in b_all if x["id"] == branch_cabang["id"])
        assert bc.get("printer_name") == "TEST_PRINTER_A"
        assert bc.get("print_header") == "TEST_HEADER_TXT"
        assert bc.get("print_footer") == "TEST_FOOTER_TXT"

        # queue/state.settings should include print_header/print_footer for this branch
        st = requests.get(f"{API}/queue/state", params={"branch_id": branch_cabang["id"]}).json()
        assert st["settings"].get("print_header") == "TEST_HEADER_TXT"
        assert st["settings"].get("print_footer") == "TEST_FOOTER_TXT"

    def test_settings_footer_text_persists(self, admin_h):
        # Read current settings to preserve
        cur = requests.get(f"{API}/settings").json()
        payload = {
            "org_name": cur.get("org_name") or "BPR Bangunarta",
            "tagline": cur.get("tagline", ""),
            "ticker_text": cur.get("ticker_text", ""),
            "promo_media": cur.get("promo_media", []),
            "primary_color": cur.get("primary_color") or "#034871",
            "logo_url": cur.get("logo_url", ""),
            "footer_text": "TEST_FOOTER_APP_ITER4",
        }
        r = requests.put(f"{API}/settings", headers=admin_h, json=payload)
        assert r.status_code == 200
        st = requests.get(f"{API}/queue/state").json()
        assert st["settings"].get("footer_text") == "TEST_FOOTER_APP_ITER4"

        # Restore original footer_text
        payload["footer_text"] = cur.get("footer_text", "")
        requests.put(f"{API}/settings", headers=admin_h, json=payload)


# ---------- Cleanup ----------
@pytest.fixture(scope="module", autouse=True)
def cleanup_at_end(admin_h, branch_cabang):
    yield
    # Reset branch printer fields set during test
    try:
        b = requests.get(f"{API}/branches").json()
        bc = next((x for x in b if x["name"] == "Kantor Cabang"), None)
        if bc and bc.get("printer_name") == "TEST_PRINTER_A":
            requests.put(f"{API}/branches/{bc['id']}", headers=admin_h, json={
                "name": bc["name"], "address": bc.get("address", ""), "active": True,
                "ticker_text": bc.get("ticker_text", ""), "promo_media": bc.get("promo_media", []),
                "printer_name": "", "print_header": "", "print_footer": "",
            })
    except Exception:
        pass
